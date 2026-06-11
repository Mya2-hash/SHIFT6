import streamlit as st
import pandas as pd
import random
from datetime import datetime, date, time
import io
import os
import holidays
import ast
from openpyxl import load_workbook
from openpyxl.styles import Font

TEMP_FILE = ".temp_backup.xlsx"
AUTH_KEY = "next_auth_token"
AUTH_VAL = "verified_2026"

# --- [비밀번호 잠금 및 새로고침 유지 설정] ---
def check_password():
    try: query_token = st.query_params.get(AUTH_KEY, "")
    except: query_token = ""

    if st.session_state.get("password_correct", False) or query_token == AUTH_VAL:
        st.session_state["password_correct"] = True
        try: st.query_params[AUTH_KEY] = AUTH_VAL
        except: pass
        return True

    def password_entered():
        if st.session_state.get("password", "") == "nextss":
            st.session_state["password_correct"] = True
            try: st.query_params[AUTH_KEY] = AUTH_VAL
            except: pass
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    st.markdown("<h2 style='text-align: center; color: #1E3A8A;'>🏢 NEXT STAFF SERVICE</h2>", unsafe_allow_html=True)
    st.text_input("社内共通パスワードを入力してください (NEXT2026)", type="password", on_change=password_entered, key="password")
    
    if st.session_state.get("password_correct") == False:
        st.error("❌ パスワードが間違っています。")
    return False

if not check_password():
    st.stop()

st.set_page_config(page_title="2026 Smart Scheduler", layout="wide", page_icon="📅")

# --- [데이터 복구 엔진 (지정출근일 포함)] ---
def apply_backup_data(xls):
    if 'Locations' in xls.sheet_names:
        df_loc = pd.read_excel(xls, sheet_name='Locations')
        loc_list = df_loc.to_dict('records')
        st.session_state['loc_count_val'] = len(loc_list)
        for i, loc in enumerate(loc_list):
            st.session_state[f"ln_{i}"] = str(loc.get('loc_name', f"LOC {i+1}"))
            st.session_state[f"lm_{i}"] = int(loc.get('loc_min', 1))
            c_days = loc.get('closed_days', [])
            if isinstance(c_days, str):
                try: c_days = ast.literal_eval(c_days)
                except: c_days = []
            st.session_state[f"lc_{i}"] = c_days

    staff_sheet = 'Staff' if 'Staff' in xls.sheet_names else ('Sheet1' if 'Sheet1' in xls.sheet_names else None)
    if staff_sheet:
        df_staff = pd.read_excel(xls, sheet_name=staff_sheet)
        staff_list = df_staff.to_dict('records')
        st.session_state['num_staff_val'] = len(staff_list)
        for i, row in enumerate(staff_list):
            s_name = row.get('pure_name', row.get('name', f"Staff{i+1}"))
            st.session_state[f"sn_{i}"] = str(s_name)
            st.session_state[f"af_{i}"] = str(row.get('affiliation', "本社"))
            st.session_state[f"to_{i}"] = int(row.get('target_off', 8))
            
            shift_str = str(row.get('shift', '09:00-18:00'))
            try:
                s_st, s_et = shift_str.split('-')
                st.session_state[f"st_{i}"] = datetime.strptime(s_st.strip(), "%H:%M").time()
                st.session_state[f"et_{i}"] = datetime.strptime(s_et.strip(), "%H:%M").time()
            except: pass

            for key_prefix, col_name in [("or", "off_list"), ("hr", "hq_list"), ("sl", "possible_locs")]:
                val = row.get(col_name, [])
                if isinstance(val, str):
                    try: val = ast.literal_eval(val)
                    except: val = []
                st.session_state[f"{key_prefix}_{i}"] = val
                
            # 지정 출근일 복구
            fl_val = row.get('fixed_locs', "{}")
            if isinstance(fl_val, str):
                try: fl_val = ast.literal_eval(fl_val)
                except: fl_val = {}
            st.session_state[f"fl_dict_{i}"] = fl_val if isinstance(fl_val, dict) else {}

def handle_upload():
    uploader = st.session_state.get('file_uploader_key')
    if uploader is not None:
        try:
            xls = pd.ExcelFile(uploader)
            apply_backup_data(xls)
            st.session_state['upload_msg'] = "success"
        except Exception as e:
            st.session_state['upload_msg'] = f"error: {e}"

if 'init_load' not in st.session_state:
    st.session_state['init_load'] = True
    if os.path.exists(TEMP_FILE):
        try:
            xls = pd.ExcelFile(TEMP_FILE)
            apply_backup_data(xls)
        except: pass

try: jp_holidays = holidays.Japan(years=2026)
except: jp_holidays = {}

weekdays_ko = ["월", "화", "수", "목", "금", "토", "일"]
weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]

lang_dict = {
    "日本語": {
        "co_name": "株式会社NEXTスタッフサービス", "author": "制作: HWANG YOUNGSEON",
        "logout": "🔒 ログアウト", "run_btn": "🚀 2026年 勤務表を生成", "download": "📥 勤務表 保存",
        "settings": "⚙️ 1. 設定", "days": "対象月", "num_staff": "全人員", 
        "loc_settings": "📍 2. 拠点設定", "loc_count": "拠点数", "loc_name": "拠点名", "loc_min": "人数", "closed_days": "休業曜日",
        "staff_settings": "👤 勤務者詳細設定", "name": "氏名", "affiliation": "所属", "hq_staff": "本社", "disp_staff": "派遣", "possible_locs": "投入可能拠点", 
        "fixed_loc": "指定出勤日", "fixed_loc_msg": "📍 拠点別の指定出勤日 (オプション)",
        "total_off": "休日数", "off_req": "希望休日", "hq_req": "本社出勤",
        "load_save": "💾 データ管理", "upload": "バックアップ", "backup_btn": "📥 PCにフルバックアップ",
        "temp_save_title": "⏳ 一時保存 (更新しても維持)", "temp_save_btn": "💾 一時保存", "temp_clear_btn": "🗑️ 初期化",
        "temp_save_ok": "✅ 一時保存完了！", "temp_clear_ok": "✅ 初期化されました！",
        "template_msg": "📁 様式アップロード", "result_title": "📊 結果", "hq_col": "★本社出勤★", "shortage": "⚠️不足", "loc_off": "X (休み)",
        "time_set": "⏰ 勤務時間", "start": "開始", "end": "終了", "msg_load": "✅ ロード成功", "msg_done": "✅ 生成完了"
    },
    "한국어": {
        "co_name": "株式会社NEXTスタッフ서비스", "author": "제작자: HWANG YOUNGSEON",
        "logout": "🔒 로그아웃", "run_btn": "🚀 통합 근무표 생성", "download": "📥 근무표 다운로드",
        "settings": "⚙️ 1. 기본 설정", "days": "대상 월", "num_staff": "전체 인원", 
        "loc_settings": "📍 2. 거점 설정", "loc_count": "거점 개수", "loc_name": "거점명", "loc_min": "인원", "closed_days": "휴무 요일",
        "staff_settings": "👤 상세 설정", "name": "성함", "affiliation": "소속", "hq_staff": "본사", "disp_staff": "파견", "possible_locs": "투입 가능", 
        "fixed_loc": "지정 출사일", "fixed_loc_msg": "📍 거점별 지정 출근일 (선택사항)",
        "total_off": "목표 휴무", "off_req": "희망 휴일", "hq_req": "본사 출사",
        "load_save": "💾 데이터 관리", "upload": "백업 업로드", "backup_btn": "📥 PC에 풀 백업하기",
        "temp_save_title": "⏳ 일시저장 (새로고침 복구)", "temp_save_btn": "💾 일시저장", "temp_clear_btn": "🗑️ 싹 지우기",
        "temp_save_ok": "✅ 서버에 일시저장 완료!", "temp_clear_ok": "✅ 초기화되었습니다!",
        "template_msg": "📁 양식 업로드", "result_title": "📊 결과", "hq_col": "★본사출사★", "shortage": "⚠️부족", "loc_off": "X (휴무)",
        "time_set": "⏰ 시간", "start": "시작", "end": "종료", "msg_load": "✅ 로드 성공", "msg_done": "✅ 완료"
    }
}

selected_lang = st.sidebar.selectbox("🌐 Language", ["日本語", "한국어"], index=0)
L = lang_dict[selected_lang]
weekdays_active = weekdays_jp if selected_lang == "日本語" else weekdays_ko

st.markdown(f"""
    <style>
    .header-bar {{ position: fixed; top: 0; left: 0; width: 100%; background: #1E3A8A; color: white; padding: 10px 20px; display: flex; justify-content: space-between; align-items: center; z-index: 1000; }}
    .stApp {{ margin-top: 60px; }}
    .stButton>button {{ width: 100%; background-color: #1E3A8A !important; color: white !important; font-weight: bold; border-radius: 8px; height: 3.5em; }}
    </style>
    <div class="header-bar"><span>🏢 {L['co_name']}</span><span>{L['author']}</span></div>
    """, unsafe_allow_html=True)

if st.sidebar.button(L["logout"]):
    try: st.query_params.clear()
    except: pass
    for key in list(st.session_state.keys()): del st.session_state[key]
    if os.path.exists(TEMP_FILE): os.remove(TEMP_FILE)
    st.rerun()

st.sidebar.header(L["load_save"])
st.sidebar.file_uploader(L["upload"], type=["xlsx"], key="file_uploader_key", on_change=handle_upload)
template_file = st.sidebar.file_uploader(L["template_msg"], type=["xlsx"])

if "target_month" not in st.session_state: st.session_state["target_month"] = 1
target_month = st.sidebar.number_input(L["days"], 1, 12, key="target_month")
days_in_month = 28 if target_month == 2 else (30 if target_month in [4,6,9,11] else 31)

def format_day(d):
    curr_d = date(2026, target_month, d)
    return f"{d} ({weekdays_active[curr_d.weekday()]})"

num_locations = st.sidebar.number_input(L["loc_count"], 1, 10, value=st.session_state.get('loc_count_val', 4), key='loc_count_val')
location_names = []; location_configs = {}
for i in range(num_locations):
    if f"ln_{i}" not in st.session_state: st.session_state[f"ln_{i}"] = f"LOC {i+1}"
    if f"lm_{i}" not in st.session_state: st.session_state[f"lm_{i}"] = 1
    if f"lc_{i}" not in st.session_state: st.session_state[f"lc_{i}"] = []
    
    with st.sidebar.expander(f"📍 {st.session_state[f'ln_{i}']}", expanded=False):
        l_name = st.text_input(L["loc_name"], key=f"ln_{i}")
        l_min = st.number_input(L["loc_min"], 0, 10, key=f"lm_{i}")
        saved_closed = st.session_state[f"lc_{i}"]
        default_idx = [weekdays_ko.index(d) for d in saved_closed if d in weekdays_ko]
        default_display = [weekdays_active[idx] for idx in default_idx]
        
        l_closed_display = st.multiselect(L["closed_days"], weekdays_active, default=default_display, key=f"lc_disp_{i}")
        l_closed = [weekdays_ko[weekdays_active.index(d)] for d in l_closed_display]
        st.session_state[f"lc_{i}"] = l_closed
        
        if l_name: location_names.append(l_name); location_configs[l_name] = {"min": l_min, "closed": l_closed}

num_staff = st.sidebar.slider(L["num_staff"], 1, 30, value=st.session_state.get('num_staff_val', 10), key='num_staff_val')
st.header(L["staff_settings"])
staff_data = []
c1, c2 = st.columns(2)
affil_options = [L["hq_staff"], L["disp_staff"]]

for i in range(num_staff):
    if f"sn_{i}" not in st.session_state: st.session_state[f"sn_{i}"] = f"Staff{i+1}"
    if f"af_{i}" not in st.session_state: st.session_state[f"af_{i}"] = affil_options[0]
    if f"to_{i}" not in st.session_state: st.session_state[f"to_{i}"] = 8
    if f"st_{i}" not in st.session_state: st.session_state[f"st_{i}"] = time(9, 0)
    if f"et_{i}" not in st.session_state: st.session_state[f"et_{i}"] = time(18, 0)
    if f"sl_{i}" not in st.session_state: st.session_state[f"sl_{i}"] = []
    if f"or_{i}" not in st.session_state: st.session_state[f"or_{i}"] = []
    if f"hr_{i}" not in st.session_state: st.session_state[f"hr_{i}"] = []
    
    if st.session_state[f"af_{i}"] not in affil_options:
        st.session_state[f"af_{i}"] = L["hq_staff"] if st.session_state[f"af_{i}"] in ["本社", "본사"] else L["disp_staff"]
    st.session_state[f"or_{i}"] = [d for d in st.session_state[f"or_{i}"] if 1 <= d <= days_in_month]
    st.session_state[f"hr_{i}"] = [d for d in st.session_state[f"hr_{i}"] if 1 <= d <= days_in_month]
    st.session_state[f"sl_{i}"] = [loc for loc in st.session_state[f"sl_{i}"] if loc in location_names]

    with (c1 if i % 2 == 0 else c2).expander(f"👤 {st.session_state[f'sn_{i}']}", expanded=False):
        col_n, col_a = st.columns([2, 1])
        s_name = col_n.text_input(L["name"], key=f"sn_{i}", label_visibility="collapsed")
        s_affil = col_a.selectbox(L["affiliation"], affil_options, key=f"af_{i}", label_visibility="collapsed")
        
        tc1, tc2 = st.columns(2)
        st_t = tc1.time_input(L["start"], key=f"st_{i}")
        et_t = tc2.time_input(L["end"], key=f"et_{i}")
        shift_str = f"{st_t.strftime('%H:%M')}-{et_t.strftime('%H:%M')}"
        
        s_locs = st.multiselect(L["possible_locs"], location_names, key=f"sl_{i}")
        
        # 거점별 지정 출근일 UI
        fixed_locs = {}
        if s_locs:
            st.caption(L["fixed_loc_msg"])
            for loc in s_locs:
                loc_idx = location_names.index(loc)
                saved_fl = st.session_state.get(f"fl_dict_{i}", {})
                default_dates = [d for d in saved_fl.get(loc, []) if 1 <= d <= days_in_month]
                f_dates = st.multiselect(f" - {loc} {L['fixed_loc']}", range(1, days_in_month + 1), default=default_dates, key=f"fl_{i}_{loc_idx}", format_func=format_day)
                if f_dates: fixed_locs[loc] = f_dates

        t_off = st.number_input(L["total_off"], 0, 20, key=f"to_{i}")
        off_l = st.multiselect(L["off_req"], range(1, days_in_month + 1), key=f"or_{i}", format_func=format_day)
        hq_l = st.multiselect(L["hq_req"], range(1, days_in_month + 1), key=f"hr_{i}", format_func=format_day)
        
        staff_data.append({
            "name": f"{s_name} ({s_affil})", "pure_name": s_name, "affiliation": s_affil,
            "shift": shift_str, "possible_locs": s_locs, "fixed_locs": fixed_locs,
            "target_off": t_off, "off_list": off_l, "hq_list": hq_l
        })

if staff_data:
    out_cfg = io.BytesIO()
    with pd.ExcelWriter(out_cfg, engine='openpyxl') as writer:
        pd.DataFrame(staff_data).drop(columns=['name']).to_excel(writer, index=False, sheet_name='Staff')
        pd.DataFrame([{"loc_name": k, "loc_min": v['min'], "closed_days": str(v['closed'])} for k, v in location_configs.items()]).to_excel(writer, index=False, sheet_name='Locations')
    st.sidebar.download_button(L["backup_btn"], out_cfg.getvalue(), f"full_backup.xlsx")
    
    st.sidebar.divider()
    st.sidebar.markdown(f"**{L['temp_save_title']}**")
    c_s1, c_s2 = st.sidebar.columns(2)
    if c_s1.button(L["temp_save_btn"]):
        with pd.ExcelWriter(TEMP_FILE, engine='openpyxl') as writer:
            pd.DataFrame(staff_data).drop(columns=['name']).to_excel(writer, index=False, sheet_name='Staff')
            pd.DataFrame([{"loc_name": k, "loc_min": v['min'], "closed_days": str(v['closed'])} for k, v in location_configs.items()]).to_excel(writer, index=False, sheet_name='Locations')
        st.sidebar.success(L["temp_save_ok"])
    if c_s2.button(L["temp_clear_btn"]):
        if os.path.exists(TEMP_FILE): os.remove(TEMP_FILE)
        try: st.query_params.clear()
        except: pass
        for key in list(st.session_state.keys()): del st.session_state[key]
        st.rerun()

if st.button(L["run_btn"]):
    schedule_results = []; holiday_list = []
    for d in range(1, days_in_month + 1):
        curr_d = date(2026, target_month, d); curr_weekday_ko = weekdays_ko[curr_d.weekday()]
        if curr_d in jp_holidays or curr_d.weekday() >= 5: holiday_list.append(d)
        res = {"Date": f"{d} ({weekdays_active[curr_d.weekday()]})"}; assigned = []
        
        # 1. 본사 최우선 배정
        hq = [f"{s['pure_name']}({s['shift']})" for s in staff_data if d in s['hq_list']]
        res[L["hq_col"]] = ", ".join(hq) if hq else "-"
        assigned.extend([s['pure_name'] for s in staff_data if d in s['hq_list']])
        
        # 2. 거점 지정 출근 우선 배정
        fixed_assignments = {loc: [] for loc in location_configs.keys()}
        for s in staff_data:
            if s['pure_name'] not in assigned and d not in s['off_list']:
                for loc, dates in s.get('fixed_locs', {}).items():
                    if d in dates and loc in location_configs:
                        fixed_assignments[loc].append(s)
                        assigned.append(s['pure_name'])
                        break
        
        # 3. 나머지 랜덤 배정
        avail = [s for s in staff_data if s['pure_name'] not in assigned and d not in s['off_list']]
        random.shuffle(avail)
        
        for loc, config in location_configs.items():
            if curr_weekday_ko in config['closed']: res[loc] = L["loc_off"]
            else:
                needed = config['min']
                already_assigned = fixed_assignments.get(loc, [])
                still_needed = max(0, needed - len(already_assigned))
                
                sel = [s for s in avail if loc in s['possible_locs']][:still_needed]
                final_loc_staff = already_assigned + sel
                res[loc] = ", ".join([f"{s['pure_name']}({s['shift']})" for s in final_loc_staff]) if final_loc_staff else L["shortage"]
                
                for s in sel: 
                    assigned.append(s['pure_name'])
                    avail.remove(s)
                    
        schedule_results.append(res)
        
    st.dataframe(pd.DataFrame(schedule_results).style.apply(lambda row: ['color: red' if int(row['Date'].split()[0]) in holiday_list else '' for _ in row], axis=1), use_container_width=True)
    out_res = io.BytesIO()
    if template_file:
        wb = load_workbook(template_file); ws = wb.active
        for i, row in enumerate(schedule_results):
            for j, val in enumerate(row.values()): ws.cell(row=2+i, column=1+j, value=val)
        wb.save(out_res)
    else:
        with pd.ExcelWriter(out_res, engine='openpyxl') as writer: pd.DataFrame(schedule_results).to_excel(writer, index=False)
    st.success(L["msg_done"])
    st.download_button(L["download"], out_res.getvalue(), f"Schedule_{target_month}.xlsx")
